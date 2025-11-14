from playwright.sync_api import sync_playwright, Page, TimeoutError as PlaywrightTimeoutError
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.styles import PatternFill
from datetime import datetime
import threading
import msvcrt
import time
import os
import re

##############################################################################################################################################
USER_PROFILES = {str(i): fr"C:\Users\Abraham\AppData\Local\Microsoft\Edge\User Data\Profile {i}"
                 for i in range(25)}
extension_path = r"C:\Users\Abraham\AppData\Local\Microsoft\Edge\User Data\Default\Extensions\ejbalbakoplchlghecdalmeeeajnimhm\12.17.2_0"
password_extension = "A6ra#am4"
##############################################################################################################################################
browser_path = r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe"
password_web = "987654"
excel_path_lxl = "Level 1 Data.xlsx"
blue_fill_ids = {
    "35486947", "67461664", "18223530", "45973639", "57048930", "86866908", "17031882", "37288964", "73543744", "78198184",
    "65004399", "103643552", "75640845", "101954643", "14678171", "27281949", "39343322", "85692081", "86107439", "63396196",
    "41082170", "106655937", "85514396", "41703102", "89042124", "106179736", "71890762", "47626091", "105614424", "45893474",
    "70130613", "105820347", "48143018", "73433791", "81839782", "95075257", "54813507", "87057540", "77197623", "10987405",
    "59183955", "33782231", "69157921", "81450781", "59476870", "65958708", "106361975", "44973221", "69012096", "10158706", 
    "53976000", "43271828", "102672580", "65638809", "56742883", "79781877", "52699993", "39877730", "101139835", "78767125",
    "86549391", "53981001", "52775461", "87166341" , "94721328", "11543748", "16274459", "34199832", "16655415", "91465361",
    "74510721", "50330841", "80710005", "17271316", "98720944", "85175531", "41553733", "91489814", "27049527", "95208058",
    "25877506", "87681422", "83287962", "102779148", "78239832", "20939718", "68208002", "86324062", "92133969", "38114401",
    "32035609", "98332790", "86103867", "76868627", "33478442", "99324487", "22758674", "16714049", "57133451", "10110426",
    "18667021", "97302176", "54524850", "107954676", "47783138", "79168537", "58396525", "30878899", "90132775", "102271095",
    "58535334", "100605313", "67464041", "27802237", "26097089", "82270325", "56448809", "107893852", "51235153", "66761043",
    "69887003", "41093405", "54703385", "15345118", "69953085", "32986752", "76969948", "93438639", "75858677", "67433260",
    "23150846", "58994845", "57482066", "27852299", "98764104", "46961966", "29980654", "23756556", "50124203", "102424673",
    "45909713", "108084137", "105986962", "45508011", "30714772", "108992828", "97139129", "82704655", "12567314", "81001327",
    "79925577", "22239846", "77083423", "10908081", "85531839", "35661215", "9558975", "73567693", "106085217","105307216",
    "96066448", "55252785", "34004640", "76621097", "36378722", "43233672", "38100577", "91782411", "14938499", "43244912",
    "51125738", "55997506", "50891136", "37033803", "75167039", "30185806", "58089979", "69455978", "83281745", "27377976",
    "72271101", "30186461", "69853768", "104348822", "71182593", "104293639", "33678480", "36803263", "17564883", "58837476",
    "18147514", "38409937", "105039583", "93291727", "69470384", "45058089", "37323232"
}
excel_path_PB = "PVC & BNB.xlsx"
excel_path_FWD = "Farming withdrawal.xlsx"
##############################################################################################################################################

def run0(profile_number: str, user_data_dir: str):
    print(f"🚀 Launching Edge with Profile {profile_number}")

    with sync_playwright() as p:
        context = p.chromium.launch_persistent_context(
            user_data_dir=user_data_dir,
            executable_path=browser_path,
            headless=False,
            args=[
                f"--disable-extensions-except={extension_path}",
                f"--load-extension={extension_path}",
            ]
        )

        Win_cl(context)

        extension_page = context.pages[0]
        safe_goto(extension_page, "chrome-extension://ejbalbakoplchlghecdalmeeeajnimhm/home.html")
        extension_page.set_viewport_size({"width": 1000, "height": 800})

        input(f"📌 Press Enter to close Profile {profile_number}")
        context.close()

##############################################################################################################################################

def run1(profile_number: str, user_data_dir: str):
    print(f"🚀 Launching Edge with Profile {profile_number}")

    with sync_playwright() as p:
        context = p.chromium.launch_persistent_context(
            user_data_dir=user_data_dir,
            executable_path=browser_path,
            headless=False,
            args=[
                f"--disable-extensions-except={extension_path}",
                f"--load-extension={extension_path}",
            ]
        )

        Win_cl(context)

        extension_page = context.pages[0]
        safe_goto(extension_page, "chrome-extension://ejbalbakoplchlghecdalmeeeajnimhm/home.html")
        extension_page.set_viewport_size({"width": 1000, "height": 800})

        site_page = context.pages[1]
        safe_goto(site_page, "https://tokenstaking.io/")
        site_page.set_viewport_size({"width": 1000, "height": 800})

        input(f"📌 Press Enter to close Profile {profile_number}")
        context.close()

##############################################################################################################################################

def run2(profile_number: str, user_data_dir: str):
    print(f"🚀 Launching Edge with Profile {profile_number}")

    with sync_playwright() as p:
        context = p.chromium.launch_persistent_context(
            user_data_dir=user_data_dir,
            executable_path=browser_path,
            headless=False,
            args=[
                f"--disable-extensions-except={extension_path}",
                f"--load-extension={extension_path}",
            ]
        )

        Win_cl(context)
        
        extension_page = context.pages[0]
        safe_goto(extension_page, "chrome-extension://ejbalbakoplchlghecdalmeeeajnimhm/home.html")
        extension_page.set_viewport_size({"width": 1000, "height": 800})

        site_page = context.pages[1]
        safe_goto(site_page, "https://tokenfarming.io/")
        site_page.set_viewport_size({"width": 1000, "height": 800})

        input(f"📌 Press Enter to close Profile {profile_number}")
        context.close()

##############################################################################################################################################

def safe_click(page: Page, selector, wait_for_after=None, action_desc=None , timeout = 10000, wait_for_load=True):
    try:
        page.wait_for_load_state("load")
        page.wait_for_selector(selector, timeout=timeout, state="visible")
        page.click(selector)
        if wait_for_load:
            page.wait_for_load_state("load")
        if wait_for_after:
            page.wait_for_selector(wait_for_after, timeout=timeout, state="visible")
        print(f"✅ {action_desc}")
        return True
    except PlaywrightTimeoutError:
        print(f"⚠️ {action_desc} - Element not found.")
        return False
    except Exception as e:
        print(f"❌ {action_desc} Failed: {e}")
        return False


def click_connect(page: Page, connect_selector: str, password_selector: str, submit_button: str, alert_button: str,
                  succ_selector: str, password_web: str, max_attempts: int = 10, inner_max_retries: int = 3):
    for attempt in range(1, max_attempts + 1):
        try:
            print(f"🔄 Attempt {attempt} started")
            
            if attempt > 1 and attempt % 3 == 0:
                print("🔄 Reloading page after 3 attempts...")
                page.reload()
                page.wait_for_load_state("networkidle")

            page.wait_for_load_state("networkidle")
            inner_retry_count = 0
            while inner_retry_count < inner_max_retries:
                try:
                    check_control_flags()
                    page.wait_for_load_state("load")
                    page.wait_for_selector(password_selector, timeout=1000, state="visible")
                    print("🔑 Password field appeared.")
                    check_control_flags()
                    page.fill(password_selector, password_web)
                    time.sleep(2)
                    page.click(submit_button)
                    print("🔐 Password entered and submit clicked.")
                    try:
                        check_control_flags()
                        page.wait_for_selector(alert_button, timeout=2000, state="visible")
                        time.sleep(2)
                        page.click(alert_button)
                        print("⚠️ Alert button appeared and clicked.")
                        check_control_flags()
                        time.sleep(2)
                        page.click(submit_button)
                        print("🔐 Submit clicked again after alert.")
                    except PlaywrightTimeoutError:
                        check_control_flags()
                        print("✅ No alert appeared after submit.")
                except PlaywrightTimeoutError:
                    check_control_flags()
                    print("ℹ️ Password field not found")
                    try:
                        check_control_flags()
                        page.wait_for_load_state("load")
                        page.wait_for_selector(succ_selector, timeout=1000, state="visible")
                        print("✅ Enter page found")
                        break
                    except PlaywrightTimeoutError:
                        check_control_flags()
                        print("ℹ️ Enter page not found")
                try:
                    check_control_flags()
                    page.wait_for_selector(connect_selector, state="visible", timeout=2000)
                    page.click(connect_selector)
                    inner_retry_count += 1
                    print(f"➡️ Clicked connect selector, inner attempt: {inner_retry_count}")
                    time.sleep(1)
                except PlaywrightTimeoutError:
                    check_control_flags()
                    print("ℹ️ Connect selector not found either.")
                    inner_retry_count += 1
                    continue
            else:
                check_control_flags()
                print(f"ℹ️ Password field did not appear after {inner_max_retries} retries.")
                continue
            return True
        except Exception as e:
            check_control_flags()
            print(f"❌ Error during attempt {attempt}: {e}")
    print("❌ Max attempts reached without success.")
    return False


def safe_goto(page: Page, url: str, retries=5, delay=2):
    for attempt in range(1, retries + 1):
        try:
            check_control_flags()
            print(f"🌐 Navigating to {url} (attempt {attempt})")
            page.goto(url)
            page.wait_for_load_state("domcontentloaded")
            print("✅ Navigation successful.")
            return page
        except Exception:
            print(f"ℹ️ Navigation error:")
            if attempt < retries:
                print(f"🔄 Retrying in {delay} seconds...")
                time.sleep(delay)
            else:
                raise
    return page


def extract_numeric_id(text):
    m = re.search(r'\b\d+\b', str(text))
    return m.group() if m else None


def get_last_cell(ws, start_row=12, max_row=63, columns=[1, 2, 3, 4, 5]):
    for row in range(max_row -1, start_row -1, -1):
        for col in columns:
            val = ws.cell(row=row, column=col).value
            if val not in (None, '', ''):
                return ws.cell(row=row, column=col).row


def save_text1_lxl(page: Page, excel_path_lxl: str, profile_number: str, max_rows: int = 60, wait_timeout: int = 30):
    try:
        if os.path.exists(excel_path_lxl):
            wb = load_workbook(excel_path_lxl)
            if f"Profile {profile_number}" in wb.sheetnames:
                ws = wb[f"Profile {profile_number}"]
            else:
                ws = wb.create_sheet(f"Profile {profile_number}")
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = f"Profile {profile_number}"

        ws.insert_rows(1, 63)

        wrap_alignment = Alignment(wrap_text=True, horizontal='right')
        wrap_alignment1 = Alignment(wrap_text=True)
        font_14 = Font(size=14)
        font_14b = Font(bold=True, size=14)

        Title1 = ws.cell(row= 1, column= 1, value="Account No.:")
        Title01 = ws.cell(row= 1, column= 4, value="Date:")        

        #staking data
        Title0 = ws.cell(row= 2, column= 1, value="Connected:")
        Title2 = ws.cell(row= 3, column= 1, value="User ID:")
        Title200 = ws.cell(row= 4, column= 1, value="Percentage %:")
        Title3 = ws.cell(row= 5, column= 1, value="Staking:")
        Title4 = ws.cell(row= 6, column= 1, value="200%:")
        Title5 = ws.cell(row= 7, column= 1, value="Reward:")
        Title6 = ws.cell(row= 8, column= 1, value="Balance:")
        Title7 = ws.cell(row= 9, column= 1, value="Daily reward:")
        Title8 = ws.cell(row= 10, column= 1, value="Next Staking:")
        Title9 = ws.cell(row= 11, column= 1, value="Token Reward:")
        #forming data
        Title00 = ws.cell(row= 2, column= 4, value="Connected:")
        Title02 = ws.cell(row= 3, column= 4, value="User ID:")
        Title03 = ws.cell(row= 5, column= 4, value="Farming:")
        Title04 = ws.cell(row= 6, column= 4, value="Reward:")
        Title05 = ws.cell(row= 7, column= 4, value="Txn. Status:")
        Title06 = ws.cell(row= 8, column= 4, value="Farming Reward:")
        Title07 = ws.cell(row= 9, column= 4, value="Early Bird Reward:")
        Title08 = ws.cell(row= 10, column= 4, value="Referral Reward:")
        Title09 = ws.cell(row= 11, column= 4, value="Total Reward:")

        Title10 = ws.cell(row= 12, column= 1, value="Token")
        Title11 = ws.cell(row= 12, column= 2, value="Staking")
        Title12 = ws.cell(row= 12, column= 4, value="Token")
        Title13 = ws.cell(row= 12, column= 5, value="Farming")
        for col in range(2, 5):
          for r in range(1, 12):
            ws.cell(row=r, column=col, value=None)

        titles_font_14b = [Title1, Title01, Title0, Title2, Title00, Title02]
        titles_font_14 = [
            Title200, Title3, Title03, Title4, Title04,
            Title5, Title05, Title6, Title06, Title7, Title07,
            Title8, Title08, Title9, Title09
        ]
        titles_bold = [Title10, Title11, Title12, Title13]

        for t in titles_font_14b:
            t.alignment = wrap_alignment
            t.font = font_14b

        for t in titles_font_14:
            t.alignment = wrap_alignment
            t.font = font_14

        for t in titles_bold:
            t.alignment = wrap_alignment1 if t in [Title11, Title13] else wrap_alignment
            t.font = Font(bold=True)

        yellow_fill = PatternFill(fill_type="solid", fgColor='FFFF00')
        blue_fill = PatternFill(fill_type="solid", fgColor='00B0F0')

        page.wait_for_load_state("load")
        for i in range(1, max_rows + 1):
            unq_id_selector = f"#tblData1 > tbody > tr:nth-child({i}) > td.text-start.px-0 strong"
            cell_a_selector = f"#tblData1 > tbody > tr:nth-child({i}) > td.text-start.px-0"
            cell_b_selector = f"#tblData1 > tbody > tr:nth-child({i}) > td.text-end"
            cell_b_fill = f"#tblData1 > tbody > tr:nth-child({i}) > td.text-end > span > strong"
            selector = f"#tblData1 > tbody > tr:nth-child({i}) > td.text-start.px-0 > small"

            unq_id = page.query_selector(unq_id_selector)
            cell_a = page.query_selector(cell_a_selector)
            cell_b = page.query_selector(cell_b_selector)
            fill_cell_b = page.query_selector(cell_b_fill)
    
            if not unq_id or not cell_a or not cell_b:
                print(f"ℹ️ Staking Level 1 Ended in Row [{i}]")
                break
            
            found_value = False    
            start_time = time.time()

            while time.time() - start_time < wait_timeout:
                element = page.query_selector(selector)
                if element is None:
                    print("❌ Selector element not found, retrying...")
                else:
                    text = element.inner_text().strip()
                    if re.match(r"Level\b", text):
                        print(f"ℹ️ Waiting, found level text: {text}")
                    else:
                        found_value = True
                        break
                time.sleep(0.5)

            if not found_value:
                print(f"ℹ️ Staking Level 1 Ended in Row [{i}] after waiting.")
                break

            unq_id_text = unq_id.inner_text().strip()
            cell_a_text = cell_a.inner_text().strip()
            cell_b_text = cell_b.inner_text().strip()
            print(f"✅ Staking Level 1 ID: {unq_id_text}")

            cell_a_cell = ws.cell(row=i + 12, column= 1, value=cell_a_text)
            cell_b_cell = ws.cell(row=i + 12, column= 2, value=cell_b_text)

            if fill_cell_b and fill_cell_b.inner_text().strip() == "0.00":
                cell_b_cell.fill = yellow_fill

            if unq_id_text in blue_fill_ids:
                cell_a_cell.fill = blue_fill

            for col in range(3, 5):
                ws.cell(row=i + 12, column=col, value=None)

            cell_a_cell.alignment = wrap_alignment
            cell_a_cell.font = font_14
            cell_b_cell.alignment = wrap_alignment1
            cell_b_cell.font = font_14

            for col_letter in ['A', 'B', 'D', 'E']:
                ws.column_dimensions[col_letter].width = 30
            for col_letter2 in ['C', 'F']:
                ws.column_dimensions[col_letter2].width = 2

        wb.save(excel_path_lxl)
        print("✅ Staking Level 1 list Saved.")
        return True
    except Exception as e:
        print(f"❌ Error during Staking Level 1 list Saving: {e}")
        return False


def save_staking_text1(page: Page, excel_path_lxl: str, profile_number: str, wait_delay=3):
    try:
        if not os.path.exists(excel_path_lxl):
            print("❌ Excel file not found to Staking First Half INFO Level 1 data.")
            return False

        wb = load_workbook(excel_path_lxl)
        if f"Profile {profile_number}" in wb.sheetnames:
            ws = wb[f"Profile {profile_number}"]
        else:
            ws = wb.create_sheet(f"Profile {profile_number}")
        
        wrap_alignment = Alignment(wrap_text=True, horizontal='left')
        font_14 = Font(size=14)
        font_14b = Font(bold=True, size=14)

        connect_selector = "#addressconnect"
        user_id_selector = "#coinrate1 > div > ul > li.copylink > span"
        staking_selector = "#tdTotalStaked"
        reward_selector = "#tdTotalReceived"

        def to_number(value):
            cleaned = value.replace(',', '').replace('%', '').strip()
            try:
                if '.' in cleaned:
                    return float(cleaned)
                else:
                    return int(cleaned)
            except ValueError:
                return value

        def wait_for_value(selector):
            val = None
            for attempt in range(15):
                if attempt > 0 and attempt % 5 == 0:
                    print(f"🔄 Reloading page on attempt {attempt}")
                    page.reload()
                    time.sleep(2)
                element = page.query_selector(selector)
                if not element:
                    print(f"⚠️ Elements not found, attempt {attempt+1}")
                    time.sleep(wait_delay)
                    continue

                text = element.inner_text().strip()
                if text == "0.00":
                    print(f"ℹ️ Waiting for the value... attempt {attempt+1}")
                    time.sleep(wait_delay)
                    continue

                num_val = to_number(text)
                val = num_val
                print(f"✅ Staking First Half Values {val} filled")
                break
            return val if val is not None else 0

        def get_percent_value(wait_delay=2, max_attempts=3):
            percent_selector = "#tblData > tbody > tr > td.text-end > small > span > span"
            percent_selector2 = "#tblData > tbody > tr:nth-child(1) > td.text-end > small > span > span"
            found = False
            per_value = 0

            for attempt in range(max_attempts):
                print(f"🔍 Checking for selector (Attempt {attempt + 1})...")

                element = page.query_selector(percent_selector)
                if element:
                    text = element.inner_text().strip()
                    if text:
                        print(f"✅ Percentage % found with text: '{text}'")
                        try:
                            per_value = to_number(text)
                            found = True
                            break
                        except Exception as e:
                            print(f"⚠️ Error converting text to number: {e}")
                    else:
                        print("ℹ️ Element found but text is empty.")
                else:
                    print("⚠️ Percentage % not found.")

                if not found:
                    element2 = page.query_selector(percent_selector2)
                    if element2:
                        text2 = element2.inner_text().strip()
                        if text2:
                            print(f"✅ Percentage % form ...2 found with text: '{text2}'")
                            try:
                                per_value = to_number(text2)
                                found = True
                                break
                            except Exception as e:
                                print(f"⚠️ Error converting text to number: {e}")
                        else:
                            print("ℹ️ Element_2 found but text is empty.")
                    else:
                        print("⚠️ Percentage %..._2 not found.")

                if not found and attempt < max_attempts - 1:
                    print(f"🔄 Reloading page... (Attempt {attempt + 2})")
                    page.reload()
                    time.sleep(wait_delay)
                time.sleep(wait_delay)
            if not found:
                print("⚠️ Value not found after all attempts.")
            return per_value

        connect_value = wait_for_value(connect_selector)
        user_id_value = wait_for_value(user_id_selector)
        per_value = get_percent_value()
        staking_value = wait_for_value(staking_selector)
        reward_value = wait_for_value(reward_selector)

        connect_cell = ws.cell(row=2, column=2, value=connect_value)        
        user_id_cell = ws.cell(row=3, column=2, value=user_id_value)
        percent_cell = ws.cell(row=4, column=2, value=per_value)
        staking_cell = ws.cell(row=5, column=2, value=staking_value)
        zoo_cell = ws.cell(row=6, column=2, value=staking_value *2 )
        reward_cell = ws.cell(row=7, column=2, value=reward_value)
        balance_cell = ws.cell(row=8, column=2, value=staking_value *2 - reward_value)

        red_fill = PatternFill(fill_type="solid", fgColor="FFFF0000")
        yellow_fill = PatternFill(fill_type="solid", fgColor="FFFFFF00")
        green_fill = PatternFill(fill_type="solid", fgColor="FF00FF00")

        if 0 <= per_value <= 35:
            percent_cell.fill = green_fill
        elif 36 <= per_value <= 38:
            percent_cell.fill = yellow_fill
        elif per_value >= 39:
            percent_cell.fill = red_fill 

        if staking_value == 0:
            staking_cell.fill = red_fill

        connect_cell.alignment = wrap_alignment
        connect_cell.font = font_14b
        user_id_cell.alignment = wrap_alignment
        user_id_cell.font = font_14b
        percent_cell.alignment = wrap_alignment
        percent_cell.font = font_14
        staking_cell.alignment = wrap_alignment
        staking_cell.font = font_14
        reward_cell.alignment = wrap_alignment
        reward_cell.font = font_14
        zoo_cell.alignment = wrap_alignment
        zoo_cell.font = font_14
        balance_cell.alignment = wrap_alignment
        balance_cell.font = font_14

        wb.save(excel_path_lxl)
        print("✅ Staking First Half INFO Level 1 data Saved.")
        return True
    except Exception:
        print(f"❌ Error during Staking First Half INFO Level 1 data Saving.")
        return False


def save_staking_text2(page: Page, excel_path_lxl: str, profile_number: str, wait_delay=3):
    try:
        if not os.path.exists(excel_path_lxl):
            print("❌ Excel file not found to Staking Second Half INFO Level 1 data.")
            return False

        wb = load_workbook(excel_path_lxl)
        if f"Profile {profile_number}" in wb.sheetnames:
            ws = wb[f"Profile {profile_number}"]
        else:
            ws = wb.create_sheet(f"Profile {profile_number}")

        wrap_alignment = Alignment(wrap_text=True, horizontal='left')
        font_14 = Font(size=14)
        font_14b = Font(bold=True, size=14)

        TR_selector = "#tdConnectedRewardBal"

        def to_number(value):
            cleaned = value.replace(',', '').replace('+ ', '+').replace('- ', '-').strip()
            try:
                if '.' in cleaned:
                    return float(cleaned)
                else:
                    return int(cleaned)
            except ValueError:
                return value

        def wait_for_value(selector):
            val = None
            for attempt in range(15):
                if attempt > 0 and attempt % 5 == 0:
                    print(f"🔄 Reloading page on attempt {attempt}")
                    page.reload()
                    time.sleep(2)
                element = page.query_selector(selector)
                if not element:
                    print(f"⚠️ Token Reward not found, attempt {attempt+1}")
                    time.sleep(wait_delay)
                    continue

                text = element.inner_text().strip()
                if text == "0.00":
                    print(f"ℹ️ Waiting for Token Reward value... attempt {attempt+1}")
                    time.sleep(wait_delay)
                    continue

                num_val = to_number(text)
                val = num_val
                print(f"✅ Token Reward Values: {val} filled")
                break
            return val if val is not None else 0

        TR_value = wait_for_value(TR_selector)
        TR_cell = ws.cell(row=11, column=2, value=TR_value)
        TR_cell.alignment = wrap_alignment
        TR_cell.font = font_14

        DR_value02 = 0
        attempt = 0
        found = False
        while attempt < 8:
            if attempt > 0 and attempt % 2 == 0:
                print(f"🔄 Reloading page on attempt {attempt + 1}")
                page.reload()
                time.sleep(2)

            DR_selector = "#tblData > tbody > tr:nth-child(1) > td.text-start.px-0 > strong"
            element = page.query_selector(DR_selector)
            if element is not None:
                DR_text = element.inner_text().strip()
                if DR_text != "":
                    try:
                        for i in range(1, 11):
                            selector01 = f"#tblData > tbody > tr:nth-child({i}) > td.text-start.px-0 > strong"
                            element01 = page.query_selector(selector01)
                            if element01:
                                text_content01 = element01.inner_text().strip()
                                if text_content01 == "Staking Reward":
                                    print("✅ Text 'Staking Reward' found")
                                    selector02 = f"#tblData > tbody > tr:nth-child({i}) > td.text-end > span > strong"
                                    element02 = page.query_selector(selector02)
                                    DR_value01 = element02.inner_text().strip()
                                    DR_value02 = to_number(DR_value01)
                                    print(f"✅ Staking Reward = {DR_value02}.")
                                    found = True 
                                    break
                                else:
                                    print("ℹ️ Text 'Staking Reward' not found")
                            else:
                                print("ℹ️ List Completed.")
                        if found:
                            break 
                    except ValueError:
                        print("ℹ️ Element not found for Staking Reward.")
            attempt += 1
            time.sleep(wait_delay)
        else:
            print("ℹ️ Staking Reward not found after all loops.")

        DR_value2 = 0
        attempt = 0
        found = False
        while attempt < 2:
            if attempt > 0:
                print(f"🔄 Reloading page on attempt {attempt + 1}")
                page.reload()
                time.sleep(2)

            DR_selector = "#tblData > tbody > tr:nth-child(1) > td.text-start.px-0 > strong"
            element = page.query_selector(DR_selector)
            if element is not None:
                DR_text = element.inner_text().strip()
                if DR_text != "":
                    try:
                        for i in range(1, 11):
                            selector1 = f"#tblData > tbody > tr:nth-child({i}) > td.text-start.px-0 > strong"
                            element1 = page.query_selector(selector1)
                            if element1:
                                text_content1 = element1.inner_text().strip()
                                if text_content1 == "Referral Reward":
                                    print("✅ Text 'Referral Reward' found")
                                    selector2 = f"#tblData > tbody > tr:nth-child({i}) > td.text-end > span > strong"
                                    element2 = page.query_selector(selector2)
                                    DR_value1 = element2.inner_text().strip()
                                    DR_value2 = to_number(DR_value1)
                                    print(f"✅ Referral Reward = {DR_value2}.")
                                    found = True 
                                    break
                                else:
                                    print(f"ℹ️ Text 'Referral Reward' not found")
                            else:
                                print("ℹ️ List Completed.")
                        if found:
                            break 
                    except ValueError:
                        print("ℹ️ Element not found for Referral Reward.")
            attempt += 1
            time.sleep(wait_delay)
        else:
            print("ℹ️ Referral Reward not found after all loops.")

        DR_value = DR_value2 + DR_value02
        print(f"➕ Referral Reward + Staking Reward = {DR_value2} + {DR_value02}")
        DR_cell = ws.cell(row=9, column=2, value=DR_value)
        print(f"✅ Daily reward value {DR_value}")

        red_fill = PatternFill(fill_type="solid", fgColor="FFFF0000")
        yellow_fill = PatternFill(fill_type="solid", fgColor="FFFFFF00")
        green_fill = PatternFill(fill_type="solid", fgColor="FF00FF00")

        balance_value = ws.cell(row=8, column=2).value
        if DR_value != 0:
            dr_num = float(DR_value)
            balance_num = float(balance_value)  
            NF_value = balance_num / dr_num
            NF_cell = ws.cell(row=10, column=2, value=NF_value)
            print(f"✅ Next Staking value {NF_value}")
            if 0 <= NF_value <= 6:
                NF_cell.fill = red_fill
            elif 7 <= NF_value <= 29:
                NF_cell.fill = yellow_fill
            elif NF_value >= 30:
                NF_cell.fill = green_fill
        else:
            NF_cell = ws.cell(row=10, column=2, value= 0 )
            print(f"✅ Next Staking value = 0")

        if DR_cell is not None:
            DR_cell.alignment = wrap_alignment
            DR_cell.font = font_14

        if NF_cell is not None:
            NF_cell.alignment = wrap_alignment
            NF_cell.font = font_14

        date_cell = ws.cell(row=1, column=5, value=datetime.now().strftime("%d/%m/%Y"))
        date_cell.alignment = wrap_alignment
        date_cell.font = font_14b

        wb.save(excel_path_lxl)
        print("✅ Staking Second Half INFO Level 1 data Saved.")
        return True
    except Exception as e:
        print(f"❌ Error during Staking Second Half INFO Level 1 data Saving.{e}")
        return False


def save_text2_lxl(page: Page, excel_path_lxl: str, profile_number: str, max_rows: int = 60, wait_timeout: int = 30):
    try:
        if not os.path.exists(excel_path_lxl):
            print("❌ Excel file not found to Farming Level 1 data.")
            return False

        wb = load_workbook(excel_path_lxl)
        if f"Profile {profile_number}" in wb.sheetnames:
            ws = wb[f"Profile {profile_number}"]
        else:
            ws = wb.create_sheet(f"Profile {profile_number}")

        id_to_row = {}
        max_search_rows = 62
        for row in range(12, max_search_rows):
            full_info = ws.cell(row=row, column=1).value
            unq_id_text = extract_numeric_id(full_info)
            if unq_id_text:
                id_to_row[str(unq_id_text).strip()] = row

        wrap_alignment1 = Alignment(wrap_text=True, horizontal='right')
        wrap_alignment = Alignment(wrap_text=True)
        font_14 = Font(size=14)

        yellow_fill = PatternFill(fill_type="solid", fgColor='FFFF00')
        blue_fill = PatternFill(fill_type="solid", fgColor='00B0F0')

        for i in range(1, max_rows + 1):
            unq_id_selector = f"#tblData1 > tbody > tr:nth-child({i}) > td.text-start.px-0 strong"
            cell_d_selector = f"#tblData1 > tbody > tr:nth-child({i}) > td.text-start.px-0"
            cell_e_selector = f"#tblData1 > tbody > tr:nth-child({i}) > td.text-end"
            cell_e_fill = f"#tblData1 > tbody > tr:nth-child({i}) > td.text-end > span > strong"
            selector = f"#tblData1 > tbody > tr:nth-child({i}) > td.text-start.px-0 > small"

            unq_id = page.query_selector(unq_id_selector)
            cell_d = page.query_selector(cell_d_selector)
            cell_e = page.query_selector(cell_e_selector)
            fill_cell_e = page.query_selector(cell_e_fill)

            if not unq_id or not cell_d or not cell_e:
                print(f"ℹ️ Farming Level 1 Ended in Row [{i}]")
                break

            found_value = False    
            start_time = time.time()

            while time.time() - start_time < wait_timeout:
                element = page.query_selector(selector)
                if element is None:
                    print("❌ Selector element not found, retrying...")
                else:
                    text = element.inner_text().strip()
                    if re.match(r"Level\b", text):
                        print(f"ℹ️ Waiting, found level text: {text}")
                    else:
                        found_value = True
                        break
                time.sleep(0.5)

            if not found_value:
                print(f"ℹ️ Farming Level 1 Ended in Row [{i}] after waiting.")
                break

            unq_id_text = unq_id.inner_text().strip()
            cell_d_text = cell_d.inner_text().strip()
            cell_e_text = cell_e.inner_text().strip()
            print(f"✅ Farming Level 1 ID: {unq_id_text}")

            if unq_id_text in id_to_row:
                row = id_to_row[unq_id_text]
                cell_d_cell = ws.cell(row=row, column=4, value=cell_d_text)
                cell_e_cell = ws.cell(row=row, column=5, value=cell_e_text)

                if fill_cell_e and fill_cell_e.inner_text().strip() == "0.00":
                    cell_e_cell.fill = yellow_fill

                if unq_id_text in blue_fill_ids:
                    cell_d_cell.fill = blue_fill

                for col in range(1, 3):
                    ws.cell(row=row, column=col, value=None)
            else:
                last_cell = get_last_cell(ws)
                ws.insert_rows(last_cell + 1)
                cell_d_cell = ws.cell(row=last_cell + 1, column=4, value=cell_d_text)
                cell_e_cell = ws.cell(row=last_cell + 1, column=5, value=cell_e_text)

                if fill_cell_e and fill_cell_e.inner_text().strip() == "0.00":
                    cell_e_cell.fill = yellow_fill

                for col in range(1, 3):
                    ws.cell(row=last_cell + 1, column=col, value=None)

            cell_d_cell.alignment = wrap_alignment1
            cell_d_cell.font = font_14
            cell_e_cell.alignment = wrap_alignment
            cell_e_cell.font = font_14

        wb.save(excel_path_lxl)
        print("✅ Farming Level 1 list Saved.")
        return True
    except Exception:
        print(f"❌ Error during Farming Level 1 list Saving.")
        return False


def save_farming_text1(page: Page, excel_path_lxl: str, profile_number: str, wait_delay=3):
    try:
        if not os.path.exists(excel_path_lxl):
            print("❌ Excel file not found to Farming First Half INFO Level 1 data.")
            return False

        wb = load_workbook(excel_path_lxl)
        if f"Profile {profile_number}" in wb.sheetnames:
            ws = wb[f"Profile {profile_number}"]
        else:
            ws = wb.create_sheet(f"Profile {profile_number}")

        wrap_alignment = Alignment(wrap_text=True, horizontal='left')
        font_14 = Font(size=14)
        font_14b = Font(bold=True, size=14)
        
        connect_selector = "#addressconnect"
        user_id_selector = "#coinrate1 > div > ul > li.copylink > span"
        farming_selector = "#tdTotalStaked"
        reward_selector = "#tdConnectedRewardBal"
        Txn_selector = "#trrequest0 > td:nth-child(4) > span > strong"

        def to_number(value):
            cleaned = value.replace(',', '').replace('+ ', '+').replace('- ', '-').strip()
            try:
                if '.' in cleaned:
                    return float(cleaned)
                else:
                    return int(cleaned)
            except ValueError:
                return value

        def wait_for_value(selector):
            val = None
            for attempt in range(15):
                if attempt > 0 and attempt % 5 == 0:
                    print(f"🔄 Reloading page on attempt {attempt}")
                    page.reload()
                    time.sleep(2)
                element = page.query_selector(selector)
                if not element:
                    print(f"⚠️ Elements not found, attempt {attempt+1}")
                    time.sleep(wait_delay)
                    continue

                text = element.inner_text().strip()
                if text == "0.00":
                    print(f"ℹ️ Waiting for the value... attempt {attempt+1}")
                    time.sleep(wait_delay)
                    continue

                num_val = to_number(text)
                val = num_val
                print(f"✅ Farming First Half Values {val} filled")
                break
            return val if val is not None else 0

        def get_Txn_value(selector, wait_delay=2, max_attempts=3):
            found = False
            per_value = 0

            for attempt in range(max_attempts):
                print(f"🔍 Checking for selector (Attempt {attempt + 1})...")

                element = page.query_selector(selector)
                if element:
                    text = element.inner_text().strip()
                    if text:
                        print(f"✅ Reward Value found with text: '{text}'")
                        try:
                            per_value = to_number(text)
                            found = True
                            break
                        except Exception as e:
                            print(f"⚠️ Error converting text to number: {e}")
                    else:
                        print("ℹ️ Element found but text is empty.")
                else:
                    print("⚠️ Reward Value not found.")

                if not found and attempt < max_attempts - 1:
                    print(f"🔄 Reloading page... (Attempt {attempt + 2})")
                    page.reload()
                    time.sleep(wait_delay)
                time.sleep(wait_delay)
            if not found:
                print("⚠️ Value not found after all attempts.")
            return per_value
        
        connect_value = wait_for_value(connect_selector)
        user_id_value = wait_for_value(user_id_selector)
        farming_value = wait_for_value(farming_selector)
        reward_value = wait_for_value(reward_selector)
        Txn_value = get_Txn_value(Txn_selector)

        connect_cell = ws.cell(row=2, column=5, value=connect_value)
        user_id_cell = ws.cell(row=3, column=5, value=user_id_value)
        farming_cell = ws.cell(row=5, column=5, value=farming_value)
        reward_cell = ws.cell(row=6, column=5, value=reward_value)
        Txn_cell = ws.cell(row=7, column=5, value=Txn_value)

        red_fill = PatternFill(fill_type="solid", fgColor="FFFF0000")

        if farming_value == 0:
            farming_cell.fill = red_fill

        connect_cell.alignment = wrap_alignment
        connect_cell.font = font_14b
        user_id_cell.alignment = wrap_alignment
        user_id_cell.font = font_14b
        farming_cell.alignment = wrap_alignment
        farming_cell.font = font_14
        reward_cell.alignment = wrap_alignment
        reward_cell.font = font_14
        Txn_cell.alignment = wrap_alignment
        Txn_cell.font = font_14

        wb.save(excel_path_lxl)
        print("✅ Farming First Half INFO Level 1 data Saved.")
        return True
    except Exception:
        print(f"❌ Error during Farming First Half INFO Level 1 data Saving.")
        return False


def save_farming_text2(page: Page, excel_path_lxl: str, profile_number: str, wait_delay=3):
    try:
        if not os.path.exists(excel_path_lxl):
            print("❌ Excel file not found to Second Half INFO Level 1 data.")
            return False

        wb = load_workbook(excel_path_lxl)
        if f"Profile {profile_number}" in wb.sheetnames:
            ws = wb[f"Profile {profile_number}"]
        else:
            ws = wb.create_sheet(f"Profile {profile_number}")

        wrap_alignment = Alignment(wrap_text=True, horizontal='left')
        font_14 = Font(size=14)

        FR_selector = "#tblData > tbody > tr:nth-child(1) > td.text-end > span > strong"
        FR_selector2 = "#tblData > tbody > tr:nth-child(1) > td.text-end > span > strong"
        EBR_selector = "#tblData1 > tbody > tr:nth-child(1) > td.text-end > span > strong"
        EBR_selector2 = "#tblData1 > tbody > tr:nth-child(1) > td.text-end > span > strong"
        RR_selector = "#tblData1 > tbody > tr:nth-child(1) > td.text-end > span > strong"
        RR_selector2 = "#tblData1 > tbody > tr:nth-child(1) > td.text-end > span > strong"

        def to_number(value):
            cleaned = value.replace(',', '').replace('+ ', '+').replace('- ', '-').strip()
            try:
                if '.' in cleaned:
                    return float(cleaned)
                else:
                    return int(cleaned)
            except ValueError:
                return value

        def get_reward_value(selector, selector2, wait_delay=2, max_attempts=3):
            found = False
            per_value = 0

            for attempt in range(max_attempts):
                print(f"🔍 Checking for selector (Attempt {attempt + 1})...")

                element = page.query_selector(selector)
                if element:
                    text = element.inner_text().strip()
                    if text:
                        print(f"✅ Reward Value found with text: '{text}'")
                        try:
                            per_value = to_number(text)
                            found = True
                            break
                        except Exception as e:
                            print(f"⚠️ Error converting text to number: {e}")
                    else:
                        print("ℹ️ Element found but text is empty.")
                else:
                    print("⚠️ Reward Value not found.")

                if not found:
                    element2 = page.query_selector(selector2)
                    if element2:
                        text2 = element2.inner_text().strip()
                        if text2:
                            print(f"✅ Reward Value_2 found with text: '{text2}'")
                            try:
                                per_value = to_number(text2)
                                found = True
                                break
                            except Exception as e:
                                print(f"⚠️ Error converting text to number: {e}")
                        else:
                            print("ℹ️ Element_2 found but text is empty.")
                    else:
                        print("⚠️ Reward Value_2 not found.")

                if not found and attempt < max_attempts - 1:
                    print(f"🔄 Reloading page... (Attempt {attempt + 2})")
                    page.reload()
                    time.sleep(wait_delay)
                time.sleep(wait_delay)
            if not found:
                print("⚠️ Value not found after all attempts.")
            return per_value

        if not logout_retry_002(page):
            raise Exception("6th click sequence failed")
        page.wait_for_load_state("load")
        time.sleep(wait_delay)
        FR_value = get_reward_value(FR_selector, FR_selector2)

        if not logout_retry_003(page):
            raise Exception("7th click sequence failed")
        page.wait_for_load_state("load")
        time.sleep(wait_delay)
        EBR_value = get_reward_value(EBR_selector, EBR_selector2)

        if not logout_retry_004(page):
            raise Exception("8th click sequence failed")
        page.wait_for_load_state("load")
        time.sleep(wait_delay)
        RR_value = get_reward_value(RR_selector, RR_selector2)

        FR_cell = ws.cell(row=8, column=5, value=FR_value)
        EBR_cell = ws.cell(row=9, column=5, value=EBR_value)
        RR_cell = ws.cell(row=10, column=5, value=RR_value)
        TR_value = FR_value + EBR_value + RR_value
        TR_cell = ws.cell(row=11, column=5, value=TR_value)
        print(f"➕ FR + EBR + RR = {FR_value} + {EBR_value} + {RR_value}")  
        print(f"✅ Total Reward value {TR_value}")

        FR_cell.alignment = wrap_alignment
        FR_cell.font = font_14
        EBR_cell.alignment = wrap_alignment
        EBR_cell.font = font_14
        RR_cell.alignment = wrap_alignment
        RR_cell.font = font_14
        TR_cell.alignment = wrap_alignment
        TR_cell.font = font_14

        wb.save(excel_path_lxl)
        print("✅ Farming Second Half INFO Level 1 data Saved.")
        return True
    except Exception as e:
        print(f"❌ Error during Farming Second Half INFO Level 1 data Saving.{e}")
        return False


pause_event = threading.Event()
skip_event = threading.Event()
stop_event = threading.Event()
def check_control_flags():
    if pause_event.is_set():
        print("⏸️ Paused, Press Enter to resume.")
    while pause_event.is_set():
        time.sleep(1)
    if stop_event.is_set():
        print("🛑 Stopping process")
        raise Exception("👤 Process stopped by user.")
    if skip_event.is_set():
        print("⏭️ Skip command received.")
        raise Exception("Skipping current account as requested.")


MAX_RETRIES = 10
def logout_retry(page: Page, account_index: int):
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            if stop_event.is_set() or skip_event.is_set():
                return False
            check_control_flags()
            if not safe_click(page,
                             "#app-content > div > div.mm-box.multichain-app-header.mm-box--margin-bottom-0.mm-box--display-flex.mm-box--align-items-center.mm-box--width-full.mm-box--background-color-background-alternative > div > div.mm-box.mm-text.mm-text--body-md.mm-text--ellipsis.mm-box--display-flex.mm-box--flex-direction-column.mm-box--align-items-center.mm-box--color-text-default > button",
                              action_desc="Open wallet list"):
                raise Exception("Open wallet list failed")
            check_control_flags()
            if not safe_click(page,
                   f"body > div.mm-modal > div:nth-child(3) > div > section > div.mm-box.multichain-account-menu-popover__list > div:nth-child({account_index}) > div > div.mm-box.multichain-account-list-item__content.mm-box--display-flex.mm-box--flex-direction-column > div.mm-box.mm-box--display-flex.mm-box--flex-direction-column > div > div.mm-box.multichain-account-list-item__account-name.mm-box--margin-inline-end-2.mm-box--display-flex.mm-box--gap-2.mm-box--align-items-center > button", 
                   action_desc=f"Select account {account_index}"):
                raise Exception(f"Select account {account_index} failed")
            return True
        except Exception as e:
            if str(e) == "👤 Process stopped by user." or str(e) == "Skipping current account as requested.":
                raise
            print(f"⚠️ Attempt {attempt} failed: {e}")
            try:
                page.goto("chrome-extension://ejbalbakoplchlghecdalmeeeajnimhm/home.html")
                page.wait_for_load_state("domcontentloaded")
                print(f"✅ Retry Page navigate in the Attempt {attempt}")
            except Exception as nav_err:
                print(f"❌ Retry Page Failed to navigate {nav_err}")
        if attempt < MAX_RETRIES:
            print(f"🔄 Retrying click logout or page (attempt {attempt + 1}/{MAX_RETRIES})...")
            time.sleep(3)
        else:
            print("❌ Max retries reached for click logout or page, aborting.")
            return False


MAX_RETRIES = 10
def logout_retry_1(page: Page):
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            if stop_event.is_set() or skip_event.is_set():
                return False
            check_control_flags()
            if not click_connect(page, "#linkConnect", "#txtPassword", "#btnLogin", "#divalert > div > div.card-body.p-3 > input", "#linklogout", password_web ):
                raise Exception("Connected and Password failed")
            page.wait_for_load_state("load")
            check_control_flags()
            if not safe_click(page, "#fireworks > div > div.container > div > div > div.menu_right_control > div.header-right.align-items-center > div.h-menu-wrap > ul > li:nth-child(4) > a", "#rbtnReferralTeam", "Click Referral Status"):
                raise Exception("Click Referral Status failed")
            check_control_flags()
            if not safe_click(page, "#rbtnReferralTeam", "#divFiltertblData1 > div", "Click Referral Team"):
                raise Exception("Click Referral Team failed")
            check_control_flags()
            if not safe_click(page, "#divFiltertblData1 > div", "#divFiltertblData1 > div > ul > li:nth-child(2)", "Click All Eligible Levels"):
                raise Exception("Click All Eligible Levels failed")
            check_control_flags()
            if not safe_click(page, "#divFiltertblData1 > div > ul > li:nth-child(2)", "", "Click Level 1"):
                raise Exception("Click Level 1 failed")
            return True
        except Exception as e:
            if str(e) == "👤 Process stopped by user." or str(e) == "Skipping current account as requested.":
                raise
            print(f"⚠️ Attempt {attempt} failed: {e}")
            if attempt in [1, 2, 4, 5, 7, 8, 10]:
                logout_clicked = safe_click(page, "#linklogout", action_desc="Click Logout on failure")
                if logout_clicked:
                    print("ℹ️ Logout clicked")
                else:
                    print("⚠️ Logout button not found")
            elif attempt in [3, 6, 9]:
                try:
                    page.goto("https://tokenstaking.io/")
                    page.wait_for_load_state("domcontentloaded")
                    print(f"✅ Retry Page navigate in the Attempt {attempt}")
                except Exception as nav_err:
                    print(f"❌ Retry Page Failed to navigate {nav_err}")
            if attempt < MAX_RETRIES:
                print(f"🔄 Retrying click sequence (attempt {attempt + 1}/{MAX_RETRIES})...")
                time.sleep(3)
            else:
                print("❌ Max retries reached for click sequence, aborting.")
                return False
  

MAX_RETRIES = 10
def logout_retry_01(page: Page):
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            if stop_event.is_set() or skip_event.is_set():
                return False
            check_control_flags()
            if not click_connect(page, "#linkConnect", "#txtPassword", "#btnLogin", "#divalert > div > div.card-body.p-3 > input", "#linklogout", password_web):
                raise Exception("Connected and Password Entered")
            page.wait_for_load_state("load")
            check_control_flags()
            if not safe_click(page, "#fireworks > div > div.container > div > div > div.menu_right_control > div.header-right.align-items-center > div.h-menu-wrap > ul > li:nth-child(2) > a", "", "Click Staking Status"):
                raise Exception("Click Staking Status failed")
            return True
        except Exception as e:
            if str(e) == "👤 Process stopped by user." or str(e) == "Skipping current account as requested.":
                raise
            print(f"⚠️ Attempt {attempt} failed: {e}")
            if attempt in [1, 2, 4, 5, 7, 8, 10]:
                logout_clicked = safe_click(page, "#linklogout", action_desc="Click Logout on failure")
                if logout_clicked:
                    print("ℹ️ Logout clicked")
                else:
                    print("⚠️ Logout button not found")
            elif attempt in [3, 6, 9]:
                try:
                    page.goto("https://tokenstaking.io/")
                    page.wait_for_load_state("domcontentloaded")
                    print(f"✅ Retry Page navigate in the Attempt {attempt}")
                except Exception as nav_err:
                    print(f"❌ Retry Page Failed to navigate {nav_err}")
            if attempt < MAX_RETRIES:
                print(f"🔄 Retrying click logout or page (attempt {attempt + 1}/{MAX_RETRIES})...")
                time.sleep(3)
            else:
                print("❌ Max retries reached for click logout or page, aborting.")
                return False


MAX_RETRIES = 10
def logout_retry_02(page: Page):
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            if stop_event.is_set() or skip_event.is_set():
                return False
            check_control_flags()
            if not click_connect(page, "#linkConnect", "#txtPassword", "#btnLogin", "#divalert > div > div.card-body.p-3 > input", "#linklogout", password_web):
                raise Exception("Connected and Password Entered")
            page.wait_for_load_state("load")
            check_control_flags()
            if not safe_click(page, "#fireworks > div > div.container > div > div > div.menu_right_control > div.header-right.align-items-center > div.h-menu-wrap > ul > li:nth-child(3) > a", "", "Click Reward Status"):
                raise Exception("Click Reward Status failed")
            return True
        except Exception as e:
            if str(e) == "👤 Process stopped by user." or str(e) == "Skipping current account as requested.":
                raise
            print(f"⚠️ Attempt {attempt} failed: {e}")
            if attempt in [1, 2, 4, 5, 7, 8, 10]:
                logout_clicked = safe_click(page, "#linklogout", action_desc="Click Logout on failure")
                if logout_clicked:
                    print("ℹ️ Logout clicked")
                else:
                    print("⚠️ Logout button not found")
            elif attempt in [3, 6, 9]:
                try:
                    page.goto("https://tokenstaking.io/")
                    page.wait_for_load_state("domcontentloaded")
                    print(f"✅ Retry Page navigate in the Attempt {attempt}")
                except Exception as nav_err:
                    print(f"❌ Retry Page Failed to navigate {nav_err}")
            if attempt < MAX_RETRIES:
                print(f"🔄 Retrying click logout or page (attempt {attempt + 1}/{MAX_RETRIES})...")
                time.sleep(3)
            else:
                print("❌ Max retries reached for click logout or page, aborting.")
                return False 
            

MAX_RETRIES = 10
def logout_retry_2(page: Page):
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            if stop_event.is_set() or skip_event.is_set():
                return False
            check_control_flags()
            if not click_connect(page, "#linkConnect", "#txtPassword", "#btnLogin", "#divalert > div > div.card-body.p-3 > input", "#linklogout", password_web):
                raise Exception("Connected and Password Entered")
            page.wait_for_load_state("load")
            check_control_flags()
            if not safe_click(page, "#form1 > header > div > div.container > div > div > div > div.header-right.align-items-center > div.h-menu-wrap > ul > li:nth-child(3) > a", "#rbtnReferralTeam", "Click Referral Status"):
                raise Exception("Click Referral Status failed")
            check_control_flags()
            if not safe_click(page, "#rbtnReferralTeam", "#divFiltertblData1 > div", "Click Referral Team"):
                raise Exception("Click Referral Team failed")
            check_control_flags()
            if not safe_click(page, "#divFiltertblData1 > div", "#divFiltertblData1 > div > ul > li:nth-child(2)", "Click All Eligible Levels"):
                raise Exception("Click All Eligible Levels failed")
            check_control_flags()
            if not safe_click(page, "#divFiltertblData1 > div > ul > li:nth-child(2)", "", "Click Level 1"):
                raise Exception("Click Level 1 failed")
            return True
        except Exception as e:
            if str(e) == "👤 Process stopped by user." or str(e) == "Skipping current account as requested.":
                raise
            print(f"⚠️ Attempt {attempt} failed: {e}")
            if attempt in [1, 2, 4, 5, 7, 8, 10]:
                logout_clicked = safe_click(page, "#linklogout", action_desc="Click Logout on failure")
                if logout_clicked:
                    print("ℹ️ Logout clicked")
                else:
                    print("⚠️ Logout button not found")
            elif attempt in [3, 6, 9]:
                try:
                    page.goto("https://tokenfarming.io/")
                    page.wait_for_load_state("domcontentloaded")
                    print(f"✅ Retry Page navigate in the Attempt {attempt}")
                except Exception as nav_err:
                    print(f"❌ Retry Page Failed to navigate {nav_err}")
            if attempt < MAX_RETRIES:
                print(f"🔄 Retrying click sequence (attempt {attempt + 1}/{MAX_RETRIES})...")
                time.sleep(3)
            else:
                print("❌ Max retries reached for click sequence, aborting.")
                return False

MAX_RETRIES = 10
def logout_retry_001(page: Page):
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            if stop_event.is_set() or skip_event.is_set():
                return False
            check_control_flags()
            if not click_connect(page, "#linkConnect", "#txtPassword", "#btnLogin", "#divalert > div > div.card-body.p-3 > input", "#linklogout", password_web):
                raise Exception("Connected and Password Entered")
            page.wait_for_load_state("load")
            check_control_flags()
            if not safe_click(page, "#form1 > header > div > div.container > div > div > div > div.header-right.align-items-center > div.h-menu-wrap > ul > li:nth-child(1) > a", "", "Click Farming Status"):
                raise Exception("Click Farming Status failed")
            return True
        except Exception as e:
            if str(e) == "👤 Process stopped by user." or str(e) == "Skipping current account as requested.":
                raise
            print(f"⚠️ Attempt {attempt} failed: {e}")
            if attempt in [1, 2, 4, 5, 7, 8, 10]:
                logout_clicked = safe_click(page, "#linklogout", action_desc="Click Logout on failure")
                if logout_clicked:
                    print("ℹ️ Logout clicked")
                else:
                    print("⚠️ Logout button not found")
            elif attempt in [3, 6, 9]:
                try:
                    page.goto("https://tokenfarming.io/")
                    page.wait_for_load_state("domcontentloaded")
                    print(f"✅ Retry Page navigate in the Attempt {attempt}")
                except Exception as nav_err:
                    print(f"❌ Retry Page Failed to navigate {nav_err}")
            if attempt < MAX_RETRIES:
                print(f"🔄 Retrying click logout or page (attempt {attempt + 1}/{MAX_RETRIES})...")
                time.sleep(3)
            else:
                print("❌ Max retries reached for click logout or page, aborting.")
                return False


MAX_RETRIES = 10
def logout_retry_002(page: Page):
    for attempt in range(1, MAX_RETRIES + 1):
        try:         
            if stop_event.is_set() or skip_event.is_set():
                return False
            check_control_flags()
            if not click_connect(page, "#linkConnect", "#txtPassword", "#btnLogin", "#divalert > div > div.card-body.p-3 > input", "#linklogout", password_web):
                raise Exception("Connected and Password Entered")
            page.wait_for_load_state("load")
            check_control_flags()
            if not safe_click(page, "#form1 > header > div > div.container > div > div > div > div.header-right.align-items-center > div.h-menu-wrap > ul > li:nth-child(4) > a", "#btnFarmingList", "Click Reward Status"):
                raise Exception("Click Reward Status failed")
            return True
        except Exception as e:
            if str(e) == "👤 Process stopped by user." or str(e) == "Skipping current account as requested.":
                raise
            print(f"⚠️ Attempt {attempt} failed: {e}")
            if attempt in [1, 2, 4, 5, 7, 8, 10]:
                logout_clicked = safe_click(page, "#linklogout", action_desc="Click Logout on failure")
                if logout_clicked:
                    print("ℹ️ Logout clicked")
                else:
                    print("⚠️ Logout button not found")
            elif attempt in [3, 6, 9]:
                try:
                    page.goto("https://tokenfarming.io/")
                    page.wait_for_load_state("domcontentloaded")
                    print(f"✅ Retry Page navigate in the Attempt {attempt}")
                except Exception as nav_err:
                    print(f"❌ Retry Page Failed to navigate {nav_err}")
            if attempt < MAX_RETRIES:
                print(f"🔄 Retrying click logout or page (attempt {attempt + 1}/{MAX_RETRIES})...")
                time.sleep(3)
            else:
                print("❌ Max retries reached for click logout or page, aborting.")
                return False
            

MAX_RETRIES = 10
def logout_retry_003(page: Page):
    for attempt in range(1, MAX_RETRIES + 1):
        try:         
            if stop_event.is_set() or skip_event.is_set():
                return False
            check_control_flags()
            if not click_connect(page, "#linkConnect", "#txtPassword", "#btnLogin", "#divalert > div > div.card-body.p-3 > input", "#linklogout", password_web):
                raise Exception("Connected and Password Entered")
            page.wait_for_load_state("load")
            check_control_flags()
            if not safe_click(page, "#form1 > header > div > div.container > div > div > div > div.header-right.align-items-center > div.h-menu-wrap > ul > li:nth-child(4) > a", "#btnEarlyBirdList", "Click Reward Status"):
                raise Exception("Click Reward Status failed")
            check_control_flags()
            if not safe_click(page, "#btnEarlyBirdList", "#btnReferralList", "Click Early Bird"):
                raise Exception("Click Early Bird failed")
            return True
        except Exception as e:
            if str(e) == "👤 Process stopped by user." or str(e) == "Skipping current account as requested.":
                raise
            print(f"⚠️ Attempt {attempt} failed: {e}")
            if attempt in [1, 2, 4, 5, 7, 8, 10]:
                logout_clicked = safe_click(page, "#linklogout", action_desc="Click Logout on failure")
                if logout_clicked:
                    print("ℹ️ Logout clicked")
                else:
                    print("⚠️ Logout button not found")
            elif attempt in [3, 6, 9]:
                try:
                    page.goto("https://tokenfarming.io/")
                    page.wait_for_load_state("domcontentloaded")
                    print(f"✅ Retry Page navigate in the Attempt {attempt}")
                except Exception as nav_err:
                    print(f"❌ Retry Page Failed to navigate {nav_err}")
            if attempt < MAX_RETRIES:
                print(f"🔄 Retrying click logout or page (attempt {attempt + 1}/{MAX_RETRIES})...")
                time.sleep(3)
            else:
                print("❌ Max retries reached for click logout or page, aborting.")
                return False
            

MAX_RETRIES = 10
def logout_retry_004(page: Page):
    for attempt in range(1, MAX_RETRIES + 1):
        try:         
            if stop_event.is_set() or skip_event.is_set():
                return False
            check_control_flags()
            if not click_connect(page, "#linkConnect", "#txtPassword", "#btnLogin", "#divalert > div > div.card-body.p-3 > input", "#linklogout", password_web):
                raise Exception("Connected and Password Entered")
            page.wait_for_load_state("load")
            check_control_flags()
            if not safe_click(page, "#form1 > header > div > div.container > div > div > div > div.header-right.align-items-center > div.h-menu-wrap > ul > li:nth-child(4) > a", "#btnReferralList", "Click Reward Status"):
                raise Exception("Click Reward Status failed")
            check_control_flags()
            if not safe_click(page, "#btnReferralList", "", "Click Referral"):
                raise Exception("Click Referral failed")
            return True
        except Exception as e:
            if str(e) == "👤 Process stopped by user." or str(e) == "Skipping current account as requested.":
                raise
            print(f"⚠️ Attempt {attempt} failed: {e}")
            if attempt in [1, 2, 4, 5, 7, 8, 10]:
                logout_clicked = safe_click(page, "#linklogout", action_desc="Click Logout on failure")
                if logout_clicked:
                    print("ℹ️ Logout clicked")
                else:
                    print("⚠️ Logout button not found")
            elif attempt in [3, 6, 9]:
                try:
                    page.goto("https://tokenfarming.io/")
                    page.wait_for_load_state("domcontentloaded")
                    print(f"✅ Retry Page navigate in the Attempt {attempt}")
                except Exception as nav_err:
                    print(f"❌ Retry Page Failed to navigate {nav_err}")
            if attempt < MAX_RETRIES:
                print(f"🔄 Retrying click logout or page (attempt {attempt + 1}/{MAX_RETRIES})...")
                time.sleep(3)
            else:
                print("❌ Max retries reached for click logout or page, aborting.")
                return False
            

def delete_empty_rows(excel_path_lxl, profile_number, max_rows=62, check_columns=[1,2,3,4,5]):
    wb = load_workbook(excel_path_lxl)
    ws = wb[f"Profile {profile_number}"] if f"Profile {profile_number}" in wb.sheetnames else wb.active
    for row in range(max_rows, 0, -1):
        is_empty = True
        for col in check_columns:
            cell = ws.cell(row=row, column=col)
            if cell.value not in (None, "", ' '):
                is_empty = False
                break
        if is_empty:
            ws.delete_rows(row)
        else:
            print(f"🚫 Stop deleting at row:{row} (not empty)")
            break

    wb.save(excel_path_lxl)
    print("✅ Done deleting trailing empty rows.")


def process_account_lxl(extension_page: Page, site_page: Page, account_index: int, profile_number: str):
    try:
        extension_page.bring_to_front()
        extension_page.wait_for_timeout(1000)  
        extension_page.wait_for_load_state("load")
        if not logout_retry(extension_page, account_index):
            raise Exception("Meta mesk process failed")

        site_page.bring_to_front()
        site_page.wait_for_timeout(1000)
        safe_goto(site_page, "https://tokenstaking.io/")

        if not logout_retry_1(site_page):
            raise Exception("1st Click sequence failed")
        site_page.wait_for_load_state("load")
        time.sleep(3)
        check_control_flags()
        save_text1_lxl(site_page, excel_path_lxl, profile_number)
        
        if not logout_retry_01(site_page):
            raise Exception("2th click sequence failed")
        site_page.wait_for_load_state("load")
        time.sleep(3)
        check_control_flags()
        save_staking_text1(site_page, excel_path_lxl, profile_number)

        if not logout_retry_02(site_page):
            raise Exception("3ed click sequence failed")
        site_page.wait_for_load_state("load")
        time.sleep(3)
        check_control_flags()
        save_staking_text2(site_page, excel_path_lxl, profile_number)

        safe_goto(site_page, "https://tokenfarming.io/")
        site_page.wait_for_load_state("load")      
        time.sleep(3)

        if not logout_retry_2(site_page):
            raise Exception("4th click sequence failed")
        site_page.wait_for_load_state("load")
        time.sleep(3)
        check_control_flags()
        save_text2_lxl(site_page, excel_path_lxl, profile_number)
 
        if not logout_retry_001(site_page):
            raise Exception("5th click sequence failed")
        site_page.wait_for_load_state("load")
        time.sleep(3)
        check_control_flags()
        save_farming_text1(site_page, excel_path_lxl, profile_number)

        check_control_flags()
        save_farming_text2(site_page, excel_path_lxl, profile_number)
        
        check_control_flags()
        wb = load_workbook(excel_path_lxl)
        if f"Profile {profile_number}" in wb.sheetnames:
            ws = wb[f"Profile {profile_number}"]
        else:
            ws = wb.create_sheet(f"Profile {profile_number}")

        Title0 = ws.cell(row= 1, column=2, value=account_index)
        Title0.alignment =  Alignment(wrap_text=True, horizontal='left')
        Title0.font = Font(bold=True, size=14)
        fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
        for col in range(1, 6):
            cell = ws.cell(row=63, column=col)
            cell.fill = fill
        wb.save(excel_path_lxl)
        time.sleep(3)
        check_control_flags()
        delete_empty_rows(excel_path_lxl, profile_number)

    except Exception as e:
        if "Skipping current account as requested." in str(e):
            print(f"⏭️ Account {account_index} skipped due to user command.")
            skip_event.clear()
            return
        elif "Process stopped by user" in str(e):
            print("🛑 Stop flag detected. Exiting process_account.")
            raise
        else:
            print(f"⚠️ Process interrupted with error: {e}")
            raise


def parse_account_input(input_str):
    accounts = set()
    parts = input_str.split(',')
    for part in parts:
        part = part.strip()
        if '-' in part:
            start, end = part.split('-')
            start, end = int(start), int(end)
            accounts.update(range(start, end + 1))
        elif part.isdigit():
            accounts.add(int(part))
        else:
            print(f"⚠️ Skipping invalid input part: {part}")
    return sorted(accounts)


def user_input_monitor():
    print("\nControls:"
          "\n [Space] = Pause"
          "\n [Enter] = Resume"
          "\n [Backspace] = Skip current task"
          "\n [Esc] = Stop all"
          "\n ")
    
    while True:
        if msvcrt.kbhit():
            key = msvcrt.getwch()
            if key == ' ':
                pause_event.set()
                print("\n⏸️ Paused.")
            elif key == '\r':
                pause_event.clear()
                print("\n▶️ Resumed.")
            elif key == '\x08':
                skip_event.set()
                print("\n⏭️ Skip requested.")
            elif key == '\x1b':
                stop_event.set()
                print("\n🛑 Stop requested.")
                break
            else:
                print(f"\n⚠️ Unknown key pressed: {repr(key)}")
        time.sleep(0.1)


def pre_process_accounts_lxl(accounts_to_process, extension_page, site_page, profile_number):
    for account_index in accounts_to_process:
        print(f"\n🔄 Starting process for account [{account_index}]")
        if stop_event.is_set():
            print("🛑 Stop detected before processing account. Exiting loop.")
            break
        try:
            process_account_lxl(extension_page, site_page, account_index, profile_number)
            print(f"✅ Completed process for account {account_index}\n")
        except Exception as e:
            err_msg = str(e)
            if "Process stopped by user" in err_msg:
                print("🛑 Stop flag detected. Exiting.")
                break
            elif skip_event.is_set():
                print(f"⏭️ Account {account_index} skipped due to user command.")
                skip_event.clear()
                continue
            else:
                print(f"❌ Error processing account {account_index}: {e}")
                continue
        time.sleep(1)


MAX_PAGES = 2
EXEMPT_URL = "chrome-extension://ejbalbakoplchlghecdalmeeeajnimhm/notification.html"
def Win_cl(context):
    def enforce_limit():
        pages = context.pages
        normal_pages = [
            p for p in pages
            if not (p.url and p.url.startswith(EXEMPT_URL))
        ]
        if len(normal_pages) > MAX_PAGES:
            print(f"⚠️ Extra pages ({len(normal_pages)}), closing extras...")
            for p in normal_pages[MAX_PAGES:]:
                try:
                    p.close()
                    print(f"🔒 Closed page: {p.url if p.url else 'about:blank'}")
                except Exception as e:
                    print(f"❌ Could not close a page: {e}")

    def on_page(page: Page):
        print(f"🆕 New page opened: {page.url if page.url else '(not loaded yet)'}")
        enforce_limit()

    def on_close(page: Page):
        print(f"❌ Page closed: {page.url if page.url else '(unknown)'}")

    context.on("page", on_page)
    context.on("close", on_close)

    enforce_limit()

    def monitor():
        while True:
            time.sleep(1.0)
            enforce_limit()
    threading.Thread(target=monitor, daemon=True).start()
    print(f"👀 Monitoring pages — keeping up to {MAX_PAGES}")


def unlock_wallet(extension_page: Page, password_extension: str):
    try:
        extension_page.wait_for_selector('#password', timeout=120000, state="visible")
        extension_page.fill('#password', password_extension)
        extension_page.click('#app-content > div > div.mm-box.main-container-wrapper > div > div > button')
        print("✅ Wallet unlocked.")
    except PlaywrightTimeoutError:
        print("ℹ️ Wallet already unlocked or password prompt not found.")
    except Exception as e:
        print(f"❌ Wallet unlock check failed: {e}")

def run3(profile_number: str, user_data_dir: str):
    user_input = input("\n📱 Enter account numbers = ")
    accounts_to_process = parse_account_input(user_input)
    print(f"💻 Accounts to process:{accounts_to_process}")
    print(f"🚀 Launching Edge with Profile {profile_number}\n")

    with sync_playwright() as p:
        context = p.chromium.launch_persistent_context(
            user_data_dir=user_data_dir,
            executable_path=browser_path,
            headless=False,
            args=[
                f"--disable-extensions-except={extension_path}",
                f"--load-extension={extension_path}",
            ]
        )

        Win_cl(context)

        extension_page = context.pages[0]
        safe_goto(extension_page, "chrome-extension://ejbalbakoplchlghecdalmeeeajnimhm/home.html")
        extension_page.set_viewport_size({"width": 1000, "height": 800})

        unlock_wallet(extension_page, password_extension)

        site_page = context.pages[1]
        site_page.set_viewport_size({"width": 1000, "height": 800})

        input_thread = threading.Thread(target=user_input_monitor, daemon=True)
        input_thread.start()

        try:
            pre_process_accounts_lxl(accounts_to_process, extension_page, site_page, profile_number)
        finally:
            print("👌 All done or stopped.")
            input(f"📌 Press Enter to close Profile {profile_number}")
            context.close()

##############################################################################################################################################

def check_bnb_balance(page: Page, threshold: float = 0.0005, wait_time: int = 60, check_interval: int = 1):
    selector = "#app-content > div > div.mm-box.main-container-wrapper > div > div > div.mm-box.mm-box--padding-top-5.mm-box--display-flex.mm-box--flex-direction-column > div.mm-box.mm-box--display-flex.mm-box--gap-4.mm-box--flex-direction-row.mm-box--width-full.mm-box--height-full > a > div.mm-box.mm-box--display-flex.mm-box--flex-direction-column.mm-box--justify-content-center.mm-box--width-full > div:nth-child(2) > p"
    QR = "#app-content > div > div.mm-box.main-container-wrapper > div > div > div.mm-box.mm-box--margin-top-4 > div > button:nth-child(5)"

    while True:
        try:
            check_control_flags()
            element = page.wait_for_selector(selector, timeout=5000, state="visible")
            text = element.text_content().strip()

            match = re.match(r"([0-9]*\.?[0-9]+)\s*BNB", text)
            if not match:
                print(f"⚠️ Could not parse balance text: {text}")
                continue

            balance = float(match.group(1))
            print(f"💰 Balance found: {balance} BNB")
            
            if balance >= threshold:
                print(f"✅ Balance {balance} is HIGHER")
                return True
            
            print(f"⚠️ Balance {balance} is LOWER, opening QR...")
            safe_click(page, QR, action_desc="QR code opened")

            check_control_flags()
            start_time = time.time()
            last_printed_balance = None
            while time.time() - start_time < wait_time:
                element = page.wait_for_selector(selector, timeout=5000, state="visible")
                text = element.text_content().strip()
                
                check_control_flags()
                match = re.match(r"([0-9]*\.?[0-9]+)\s*BNB", text)
                if match:
                    balance = float(match.group(1))
                    check_control_flags()
                    if last_printed_balance != balance:
                        print(f"💰 {balance} BNB waiting for the change...")
                        last_printed_balance = balance
                    
                    check_control_flags()
                    if balance >= threshold:
                        print(f"✅ Balance {balance} is HIGHER")
                        return True

                time.sleep(check_interval)

            print("⏳ Balance did not reach, reloading page...")
            page.reload(wait_until="load")

        except PlaywrightTimeoutError:
            print("❌ Balance element not found, retrying...")
            time.sleep(check_interval)


def process_account_BNB(extension_page: Page, account_index: int):
    try:
        extension_page.wait_for_load_state("load")
        if not logout_retry(extension_page, account_index):
            raise Exception("Meta mesk process failed")

        check_control_flags()
        extension_page.wait_for_load_state("load")
        check_bnb_balance(extension_page)

    except Exception as e:
        if "Skipping current account as requested." in str(e):
            print(f"⏭️ Account {account_index} skipped due to user command.")
            skip_event.clear()
            try:
                extension_page.reload(wait_until="load")
                print("🔄 Page reloaded after skip.")
            except Exception as reload_err:
                print(f"⚠️ Failed to reload page after skip: {reload_err}")
            return
        elif "Process stopped by user" in str(e):
            print("🛑 Stop flag detected. Exiting process_account.")
            raise
        else:
            print(f"⚠️ Process interrupted with error: {e}")
            raise

def pre_process_accounts_BNB(accounts_to_process, extension_page):
    for account_index in accounts_to_process:
        print(f"\n🔄 Starting process for account [{account_index}]")
        if stop_event.is_set():
            print("🛑 Stop detected before processing account. Exiting loop.")
            break
        try:
            process_account_BNB(extension_page, account_index)
            print(f"✅ Completed process for account {account_index}\n")
        except Exception as e:
            err_msg = str(e)
            if "Process stopped by user" in err_msg:
                print("🛑 Stop flag detected. Exiting.")
                break
            elif skip_event.is_set():
                print(f"⏭️ Account {account_index} skipped due to user command.")
                skip_event.clear()
                continue
            else:
                print(f"❌ Error processing account {account_index}: {e}")
                continue
        time.sleep(1)

def click_bnb(page: Page, max_wait: int = 30, check_interval: int = 2):
    while True:
        start_time = time.time()
        check_control_flags()
        while time.time() - start_time < max_wait:
            for i in range(1, 11):
                selector = f"#app-content > div > div.mm-box.main-container-wrapper > div > div > div > div:nth-child(3) > div > div > div > div:nth-child({i}) > a > div.mm-box.mm-box--display-flex.mm-box--flex-direction-column.mm-box--justify-content-center.mm-box--width-full > div:nth-child(1) > p"
                try:
                    element = page.wait_for_selector(selector, timeout=2000, state="visible")
                    text = element.text_content().strip() if element else ""
                    if text == "BNB":
                        check_control_flags()
                        safe_click(page, selector, action_desc=f"Click BNB at index {i}")
                        print(f"✅ BNB found and clicked at index {i}")
                        return True
                except PlaywrightTimeoutError:
                    continue

            time.sleep(check_interval)

        print("⏳ BNB not found in 30s, reloading page...")
        page.reload(wait_until="load")


def run4(profile_number: str, user_data_dir: str):
    user_input = input("\n📱 Enter account numbers = ")
    accounts_to_process = parse_account_input(user_input)
    print(f"💻 Accounts to process:{accounts_to_process}")
    print(f"🚀 Launching Edge with Profile {profile_number}\n")

    with sync_playwright() as p:
        context = p.chromium.launch_persistent_context(
            user_data_dir=user_data_dir,
            executable_path=browser_path,
            headless=False,
            args=[
                f"--disable-extensions-except={extension_path}",
                f"--load-extension={extension_path}",
            ]
        )

        Win_cl(context)

        extension_page = context.pages[0]
        safe_goto(extension_page, f"chrome-extension://ejbalbakoplchlghecdalmeeeajnimhm/home.html")
        extension_page.set_viewport_size({"width": 1000, "height": 800})
        extension_page.bring_to_front()

        unlock_wallet(extension_page, password_extension)

        site_page = context.pages[1]
        site_page.set_viewport_size({"width": 1000, "height": 800})

        input_thread = threading.Thread(target=user_input_monitor, daemon=True)
        input_thread.start()

        click_bnb(extension_page)
        extension_page.wait_for_load_state("load")

        try:
            pre_process_accounts_BNB(accounts_to_process, extension_page)
        finally:
            print("👌 All done or stopped.")
            input(f"📌 Press Enter to close Profile {profile_number}")
            context.close()

##############################################################################################################################################

def save_text1(excel_path: str, profile_number: str):
    try:
        if os.path.exists(excel_path):
            wb = load_workbook(excel_path)
            if f"Profile {profile_number}" in wb.sheetnames:
                ws = wb[f"Profile {profile_number}"]
            else:
                ws = wb.create_sheet(f"Profile {profile_number}")
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = f"Profile {profile_number}"

        ws.insert_rows(1, 1)

        wrap_alignment = Alignment(wrap_text=True, horizontal='right')
        wrap_alignment2 = Alignment(wrap_text=True, horizontal='left')
        font_14b = Font(bold=True, size=14)
        
        Title1 = ws.cell(row= 1, column=1, value="Date:")
        date_cell = ws.cell(row=1, column=2, value=datetime.now().strftime("%d/%m/%Y"))
        Title01 = ws.cell(row= 1, column=3, value="Address ID.:")
        Title0 = ws.cell(row= 1, column=5, value="Account No.:")

        for col in range(2, 5):
            ws.cell(row=1, column=col, value=None)

        Title1.alignment = wrap_alignment
        Title1.font = font_14b
        date_cell.alignment = wrap_alignment2
        date_cell.font = font_14b
        Title01.alignment = wrap_alignment
        Title01.font = font_14b
        Title0.alignment = wrap_alignment
        Title0.font = font_14b

        for col_letter in ['A']:
            ws.column_dimensions[col_letter].width = 8
        for col_letter2 in ['B', 'C', 'E']:
            ws.column_dimensions[col_letter2].width = 18

        wb.save(excel_path)
        print("✅ INFO Details saved")
        return True
    except Exception as e:
        print(f"❌ Error during INFO Details,{e}")
        return False

def save_text2(page: Page, excel_path: str, profile_number: str):
    try:
        if not os.path.exists(excel_path):
            print("❌ Excel file not found")
            return False

        wb = load_workbook(excel_path)
        if f"Profile {profile_number}" in wb.sheetnames:
            ws = wb[f"Profile {profile_number}"]
        else:
            ws = wb.create_sheet(f"Profile {profile_number}")

        wrap_alignment1 = Alignment(wrap_text=True)
        font_14 = Font(size=14)
        
        check_control_flags()
        while True:
            start_time = time.time()
            while time.time() - start_time < 30:
                for i in range(1, 11):
                    selector = f"#app-content > div > div.mm-box.main-container-wrapper > div > div > div > div:nth-child(3) > div > div > div > div:nth-child({i}) > a > div.mm-box.mm-box--display-flex.mm-box--flex-direction-column.mm-box--justify-content-center.mm-box--width-full > div:nth-child(1) > p"
                    try:
                        check_control_flags()
                        element = page.query_selector(selector)
                        if not element:
                            continue

                        check_control_flags()
                        text = element.inner_text().strip()
                        if text == "PVC META":
                            value_selector = f"#app-content > div > div.mm-box.main-container-wrapper > div > div > div > div:nth-child(3) > div > div > div > div:nth-child({i}) > a > div.mm-box.mm-box--display-flex.mm-box--flex-direction-column.mm-box--justify-content-center.mm-box--width-full > div:nth-child(2) > p"
                            PVC_data = page.query_selector(value_selector)

                            check_control_flags()
                            if PVC_data:
                                PVC_text = PVC_data.inner_text().strip()
                                cell = ws.cell(row=1, column=7, value=PVC_text)
                                cell.alignment = wrap_alignment1
                                cell.font = font_14
                                ws.column_dimensions['G'].width = 20

                                wb.save(excel_path)
                                print("✅ Saved PVC Meta:", PVC_text)
                                return True
                    except PlaywrightTimeoutError:
                        continue
                time.sleep(1)

            print("⏳ PVC Meta not found in 30s, reloading page...")
            page.reload(wait_until="load")

    except Exception as e:
        print(f"❌ Error during saving PVC Meta Data: {e}")
        return False


def save_text3(page: Page, excel_path: str, profile_number: str):
    try:
        if not os.path.exists(excel_path):
            print("❌ Excel file not found")
            return False

        wb = load_workbook(excel_path)
        if f"Profile {profile_number}" in wb.sheetnames:
            ws = wb[f"Profile {profile_number}"]
        else:
            ws = wb.create_sheet(f"Profile {profile_number}")

        wrap_alignment1 = Alignment(wrap_text=True)
        font_14 = Font(size=14)

        check_control_flags()
        while True:
            start_time = time.time()
            while time.time() - start_time < 30:
                for i in range(1, 11):
                    selector = f"#app-content > div > div.mm-box.main-container-wrapper > div > div > div > div:nth-child(3) > div > div > div > div:nth-child({i}) > a > div.mm-box.mm-box--display-flex.mm-box--flex-direction-column.mm-box--justify-content-center.mm-box--width-full > div:nth-child(1) > p"
                    try:
                        check_control_flags()
                        element = page.query_selector(selector)
                        if not element:
                            continue

                        check_control_flags()
                        text = element.inner_text().strip()
                        if text == "BNB":
                            value_selector = f"#app-content > div > div.mm-box.main-container-wrapper > div > div > div > div:nth-child(3) > div > div > div > div:nth-child({i}) > a > div.mm-box.mm-box--display-flex.mm-box--flex-direction-column.mm-box--justify-content-center.mm-box--width-full > div:nth-child(2) > p"
                            BNB_data = page.query_selector(value_selector)

                            check_control_flags()
                            if BNB_data:
                                BNB_text = BNB_data.inner_text().strip()
                                cell = ws.cell(row=1, column=8, value=BNB_text)
                                cell.alignment = wrap_alignment1
                                cell.font = font_14
                                ws.column_dimensions['H'].width = 18

                                wb.save(excel_path)
                                print("✅ Saved BNB:", BNB_text)
                                return True
                    except PlaywrightTimeoutError:
                        continue
                time.sleep(1)

            print("⏳ BNB not found in 30s, reloading page...")
            page.reload(wait_until="load")

    except Exception as e:
        print(f"❌ Error during saving BNB Data: {e}")
        return False


def process_account_PB(extension_page: Page, account_index: int, profile_number):
    try:
        extension_page.bring_to_front()
        extension_page.wait_for_timeout(1000)  
        extension_page.wait_for_load_state("load")
        if not logout_retry(extension_page, account_index):
            raise Exception("Meta mesk process failed")

        extension_page.wait_for_load_state("load")
        extension_page.wait_for_timeout(1000)
        check_control_flags()
        save_text1(excel_path_PB, profile_number)

        check_control_flags()
        save_text2(extension_page, excel_path_PB, profile_number)

        check_control_flags()
        save_text3(extension_page, excel_path_PB, profile_number)

        check_control_flags()
        wb = load_workbook(excel_path_PB)
        if f"Profile {profile_number}" in wb.sheetnames:
            ws = wb[f"Profile {profile_number}"]
        else:
            ws = wb.create_sheet(f"Profile {profile_number}")
        acc_no = ws.cell(row= 1, column=6, value=account_index)
        acc_no.alignment =  Alignment(wrap_text=True, horizontal='left')
        acc_no.font = Font(bold=True, size=14)
        ws.column_dimensions['F'].width = 8

        add_sel = "#app-content > div > div.mm-box.multichain-app-header.mm-box--margin-bottom-0.mm-box--display-flex.mm-box--align-items-center.mm-box--width-full.mm-box--background-color-background-alternative > div > div.mm-box.mm-text.mm-text--body-md.mm-text--ellipsis.mm-box--display-flex.mm-box--flex-direction-column.mm-box--align-items-center.mm-box--color-text-default > div > div > button > span.mm-box.mm-text.mm-text--inherit.mm-text--ellipsis.mm-box--display-flex.mm-box--gap-2.mm-box--align-items-center.mm-box--color-text-default > span"
        add_data = extension_page.query_selector(add_sel)
        add_text = add_data.inner_text().strip()
        add_txt = ws.cell(row= 1, column=4, value=add_text)
        add_txt.alignment =  Alignment(wrap_text=True, horizontal='left')
        add_txt.font = Font(bold=True, size=14)
        ws.column_dimensions['D'].width = 24
        wb.save(excel_path_PB)

    except Exception as e:
        if "Skipping current account as requested." in str(e):
            print(f"⏭️ Account {account_index} skipped due to user command.")
            skip_event.clear()
            return
        elif "Process stopped by user" in str(e):
            print("🛑 Stop flag detected. Exiting process_account.")
            raise
        else:
            print(f"⚠️ Process interrupted with error: {e}")
            raise


def pre_process_accounts_PB(accounts_to_process, extension_page, profile_number):
    for account_index in accounts_to_process:
        print(f"\n🔄 Starting process for account [{account_index}]")
        if stop_event.is_set():
            print("🛑 Stop detected before processing account. Exiting loop.")
            break
        try:
            process_account_PB(extension_page, account_index, profile_number)
            print(f"✅ Completed process for account {account_index}\n")
        except Exception as e:
            err_msg = str(e)
            if "Process stopped by user" in err_msg:
                print("🛑 Stop flag detected. Exiting.")
                break
            elif skip_event.is_set():
                print(f"⏭️ Account {account_index} skipped due to user command.")
                skip_event.clear()
                continue
            else:
                print(f"❌ Error processing account {account_index}: {e}")
                continue
        time.sleep(1)


def run5(profile_number: str, user_data_dir: str):
    user_input = input("\n📱 Enter account numbers = ")
    accounts_to_process = parse_account_input(user_input)
    print(f"💻 Accounts to process:{accounts_to_process}")
    print(f"🚀 Launching Edge with Profile {profile_number}\n")

    with sync_playwright() as p:
        context = p.chromium.launch_persistent_context(
            user_data_dir=user_data_dir,
            executable_path=browser_path,
            headless=False,
            args=[
                f"--disable-extensions-except={extension_path}",
                f"--load-extension={extension_path}",
            ]
        )

        Win_cl(context)

        extension_page = context.pages[0]
        safe_goto(extension_page, f"chrome-extension://ejbalbakoplchlghecdalmeeeajnimhm/home.html")
        extension_page.set_viewport_size({"width": 1000, "height": 800})

        unlock_wallet(extension_page, password_extension)

        site_page = context.pages[1]
        site_page.set_viewport_size({"width": 1000, "height": 800})

        input_thread = threading.Thread(target=user_input_monitor, daemon=True)
        input_thread.start()

        try:
            pre_process_accounts_PB(accounts_to_process, extension_page, profile_number)
        finally:
            print("👌 All done or stopped.")
            input(f"📌 Press Enter to close Profile {profile_number}")
            context.close()

##############################################################################################################################################

def save_text1_FWD(page: Page, excel_path: str, profile_number: str):
    try:
        if os.path.exists(excel_path):
            wb = load_workbook(excel_path)
            if f"Profile {profile_number}" in wb.sheetnames:
                ws = wb[f"Profile {profile_number}"]
            else:
                ws = wb.create_sheet(f"Profile {profile_number}")
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = f"Profile {profile_number}"

        ws.insert_rows(1, 1)

        def to_number(value):
            cleaned = value.replace(' ', '').strip()
            try:
                if '.' in cleaned:
                    return float(cleaned)
                else:
                    return int(cleaned)
            except ValueError:
                return value
            
        wrap_alignment = Alignment(wrap_text=True, horizontal='right')
        wrap_alignment2 = Alignment(wrap_text=True, horizontal='left')
        font_14b = Font(bold=True, size=14)
        
        user_id_selector = "#coinrate1 > div > ul > li.copylink > span"
        connect_selector = "#addressconnect"

        Title1 = ws.cell(row= 1, column=1, value="Date:")
        date_cell = ws.cell(row=1, column=2, value=datetime.now().strftime("%d/%m/%Y"))
        ID_cell = ws.cell(row= 1, column=3, value="User ID.:")
        Title01 = ws.cell(row= 1, column=5, value="Address ID.:")
        Title0 = ws.cell(row= 1, column=7, value="Account No.:")

        user_id = page.wait_for_selector(user_id_selector)
        con_sel = page.wait_for_selector(connect_selector)

        user_id_value = user_id.inner_text().strip()
        user_id_nuber = to_number(user_id_value)
        connect_value = con_sel.inner_text().strip()

        user_id_cell = ws.cell(row=1, column=4, value=user_id_nuber)
        connect_cell = ws.cell(row=1, column=6, value=connect_value)        

        Title1.alignment = wrap_alignment
        Title1.font = font_14b
        date_cell.alignment = wrap_alignment2
        date_cell.font = font_14b
        ID_cell.alignment = wrap_alignment
        ID_cell.font = font_14b
        Title01.alignment = wrap_alignment
        Title01.font = font_14b
        Title0.alignment = wrap_alignment
        Title0.font = font_14b
        connect_cell.alignment = wrap_alignment2
        connect_cell.font = font_14b
        user_id_cell.alignment = wrap_alignment2
        user_id_cell.font = font_14b

        for col_letter in ['A']:
            ws.column_dimensions[col_letter].width = 8
        for col_letter1 in ['c']:
            ws.column_dimensions[col_letter1].width = 14
        for col_letter2 in ['B', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col_letter2].width = 18
        for col_letter3 in ['I', 'K', 'J']:
            ws.column_dimensions[col_letter3].width = 32


        wb.save(excel_path)
        print("✅ INFO Details saved")
        return True
    except Exception as e:
        print(f"❌ Error during INFO Details,{e}")
        return False


MAX_RETRIES = 10
def logout_retry_FWD(page: Page):
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            if stop_event.is_set() or skip_event.is_set():
                return False
            check_control_flags()
            if not click_connect(page, "#linkConnect", "#txtPassword", "#btnLogin", "#divalert > div > div.card-body.p-3 > input", "#linklogout", password_web):
                raise Exception("Connected and Password Entered")
            check_control_flags()
            if not safe_click(page, "#form1 > header > div > div.container > div > div > div > div.header-right.align-items-center > div.h-menu-wrap > ul > li:nth-child(1) > a", "#txtWithdrawalAmount", "Click Farming Status"):
                raise Exception("Click Farming Status failed")
            return True
        except Exception as e:
            if str(e) == "👤 Process stopped by user." or str(e) == "Skipping current account as requested.":
                raise
            print(f"⚠️ Attempt {attempt} failed: {e}")
            if attempt in [1, 2, 4, 5, 7, 8, 10]:
                logout_clicked = safe_click(page, "#linklogout", action_desc="Click Logout on failure")
                if logout_clicked:
                    print("ℹ️ Logout clicked")
                else:
                    print("⚠️ Logout button not found") 
            elif attempt in [3, 6, 9]:
                try:
                    page.goto("https://tokenfarming.io/")
                    page.wait_for_load_state("domcontentloaded")
                    print(f"✅ Retry Page navigate in the Attempt {attempt}")
                except Exception as nav_err:
                    print(f"❌ Retry Page Failed to navigate {nav_err}")
            if attempt < MAX_RETRIES:
                print(f"🔄 Retrying click sequence (attempt {attempt + 1}/{MAX_RETRIES})...")
                time.sleep(3)
            else:
                print("❌ Max retries reached for click sequence, aborting.")
                return False
            

def wait_for_selector_with_timeout(page, selector, timeout):
    start_time = time.time()
    while time.time() - start_time < timeout:
        try:
            check_control_flags()
            if page.locator(selector).is_visible():
                print(f"ℹ️ Selector found")
                return True
        except Exception:
            pass
        time.sleep(1)
    print(f"⏰ Timeout waiting")
    return False


def wait_for_selector_with_timeout2(page, selector1, selector2, timeout):
    start_time = time.time()
    while time.time() - start_time < timeout:
        try:
            check_control_flags()
            if selector1 and page.locator(selector1).is_visible():
                print("ℹ️ Password Fill found")
                return selector1
            check_control_flags()
            if selector2 and page.locator(selector2).is_visible():
                print("ℹ️ Alert Message found")
                return selector2
        except Exception:
            pass
        time.sleep(1)
    print("⏰ Timeout waiting")
    return None


def process_account_FWD(extension_page: Page, site_page: Page, account_index: int, profile_number: str):
    try:
        extension_page.bring_to_front()
        extension_page.wait_for_timeout(1000)  
        extension_page.wait_for_load_state("load")
        if not logout_retry(extension_page, account_index):
            raise Exception("Meta mesk process failed")

        safe_goto(site_page, "https://tokenfarming.io/")
        site_page.bring_to_front()
        site_page.wait_for_load_state("load")      
        
        if not logout_retry_FWD(site_page):
            raise Exception("1st click sequence failed")
        site_page.wait_for_load_state("load")
        time.sleep(3)
        check_control_flags()
        save_text1_FWD(site_page, excel_path_FWD, profile_number)
        
        check_control_flags()
        wb = load_workbook(excel_path_FWD)
        if f"Profile {profile_number}" in wb.sheetnames:
            ws = wb[f"Profile {profile_number}"]
        else:
            ws = wb.create_sheet(f"Profile {profile_number}")
        acc_no = ws.cell(row= 1, column=8, value=account_index)
        acc_no.alignment =  Alignment(wrap_text=True, horizontal='left')
        acc_no.font = Font(bold=True, size=14)
        ws.column_dimensions['H'].width = 8
        wb.save(excel_path_FWD)

        def to_number(value):
            cleaned = value.replace(',', '').replace(' ', '').strip()
            try:
                if '.' in cleaned:
                    return float(cleaned)
                else:
                    return int(cleaned)
            except ValueError:
                return value

        def wait_for_value(selector):
            val = None
            for attempt in range(15):
                if attempt > 0 and attempt % 5 == 0:
                    print(f"🔄 Reloading page on attempt {attempt}")
                    site_page.reload()
                    time.sleep(2)
                element = site_page.query_selector(selector)
                if not element:
                    print(f"⚠️ Reward Status not found, attempt {attempt+1}")
                    time.sleep(2)
                    continue

                text = element.inner_text().strip()
                if text == "0.00":
                    print(f"ℹ️ Waiting for the Reward Status... attempt {attempt+1}")
                    time.sleep(2)
                    continue

                num_val = to_number(text)
                val = num_val
                print(f"✅ Reward Status {val} found")
                break
            return val if val is not None else 0

        red_fill = PatternFill(fill_type="solid", fgColor="FFFF0000")
        yellow_fill = PatternFill(fill_type="solid", fgColor="FFFFFF00")
        green_fill = PatternFill(fill_type="solid", fgColor="FF00FF00")
        check_control_flags()

        check_RS = "#tdConnectedRewardBal"
        RS_Value = wait_for_value(check_RS)
        if RS_Value >= 100:
            print("✅ Reward Status is more then 100")
            RS_txt = ws.cell(row= 1, column=9, value=f"✅ Reward Status: {RS_Value}")
            RS_txt.font = Font(size=14)
            RS_txt.fill = green_fill
            wb.save(excel_path_FWD)
            check_control_flags()

            check_PF = "#txtWithdrawalOTP"
            check_AT = "#alerttext"
            print("⏳ Waiting up for Alert or Password Fill")
            found = wait_for_selector_with_timeout2(site_page, check_PF, check_AT, timeout=300)
            if found == check_PF:
                print("✅ Password field found")
                check_control_flags()
                time.sleep(1)
                site_page.fill("#txtWithdrawalOTP", password_web)
                print("✅ Withdrawal password Filled")
                check_control_flags()
                safe_click(site_page, "#btnWithdrawal", "", "Click Submit")
                check_control_flags()

                check_P2P = "#divConfirmRedeem > div > div.card-body.p-3"
                time.sleep(1)
                site_page.wait_for_load_state("load")
                try:
                    P2P_Sell = site_page.wait_for_selector(check_P2P, timeout=5000, state="visible")
                    print("✅ Your P2P Offer appeared")
                    P2P_text = P2P_Sell.inner_text().strip()
                    P2P_txt = ws.cell(row= 1, column=10, value= f"✅ Your P2P Offer\n{P2P_text}")
                    P2P_txt.font = Font(size=14)
                    P2P_txt.fill = green_fill
                    P2P_txt.alignment = Alignment(wrap_text=True)
                    wb.save(excel_path_FWD)
                    check_control_flags()
                    safe_click(site_page, "#btnConfirmRedeem", "", "Click Confirm")
                    check_control_flags()

                    time.sleep(1)
                    site_page.wait_for_load_state("load")
                    found = wait_for_selector_with_timeout(site_page, check_AT, timeout=5)
                    if not found:
                        print("✅ No Alert message found — proceeding with transaction.")
                        check_control_flags()
                        check_TR = "#divRedeemMsg > div > div"
                        found = wait_for_selector_with_timeout(site_page, check_TR, timeout=600)
                        if not found:
                            print("❌ Transaction request message not found")
                            ATPF_txt = ws.cell(row= 1, column=11, value= "❌ Transaction request message not found")
                            ATPF_txt.font = Font(size=14)
                            ATPF_txt.alignment = Alignment(wrap_text=True)
                            ATPF_txt.fill = red_fill
                            wb.save(excel_path_FWD)
                            check_control_flags()

                        else:
                            print("✅ Transaction request message found")
                            TR_Sell = site_page.wait_for_selector(check_TR, timeout=5000, state="visible")
                            TR_text = TR_Sell.inner_text().strip()
                            TR_txt = ws.cell(row= 1, column=11, value= TR_text)
                            TR_txt.font = Font(size=14)
                            TR_txt.fill = green_fill
                            TR_txt.alignment = Alignment(wrap_text=True)
                            wb.save(excel_path_FWD)
                            check_control_flags()
                    else:
                        AT_Sell = site_page.wait_for_selector(check_AT, timeout=5000, state="visible")
                        print("⚠️ Alert appeared! after P2P Offer")
                        AT_text = AT_Sell.inner_text().strip()
                        AT_txt = ws.cell(row= 1, column=11, value= AT_text)
                        AT_txt.font = Font(size=14)
                        AT_txt.fill = yellow_fill
                        AT_txt.alignment = Alignment(wrap_text=True)
                        wb.save(excel_path_FWD)
                        check_control_flags()

                except TimeoutError:
                    print("⚠️ Your P2P Offer not found")
                    P2P_txt = ws.cell(row= 1, column=10, value="⚠️ Your P2P Offer not found")
                    P2P_txt.font = Font(size=14)
                    P2P_txt.alignment = Alignment(wrap_text=True)
                    P2P_txt.fill = red_fill
                    wb.save(excel_path_FWD)
                    check_control_flags()

            elif found == check_AT:
                ATP_Sell = site_page.wait_for_selector(check_AT, timeout=5000, state="visible")
                print("⚠️ Alert appeared! before Password Fill")
                ATP_text = ATP_Sell.inner_text().strip()
                ATP_txt = ws.cell(row= 1, column=10, value= ATP_text)
                ATP_txt.font = Font(size=14)
                ATP_txt.alignment = Alignment(wrap_text=True)
                ATP_txt.fill = red_fill
                wb.save(excel_path_FWD)
                check_control_flags()

            else:
                print("❌ Alert or Password Fill not found in time.")
                ATPF_txt = ws.cell(row= 1, column=10, value= "❌ Alert or Password Fill not found in time")
                ATPF_txt.font = Font(size=14)
                ATPF_txt.alignment = Alignment(wrap_text=True)
                ATPF_txt.fill = red_fill
                wb.save(excel_path_FWD)
                check_control_flags()
        else:
            print("⚠️ Reward Status is less then 100")
            RS_txt = ws.cell(row= 1, column=9, value=f"⚠️ Reward S is less: {RS_Value}")
            RS_txt.font = Font(size=14)
            RS_txt.alignment = Alignment(wrap_text=True)
            RS_txt.fill = red_fill
            wb.save(excel_path_FWD)
            check_control_flags()

    except Exception as e:
        if "Skipping current account as requested." in str(e):
            print(f"⏭️ Account {account_index} skipped due to user command.")
            skip_event.clear()
            return
        elif "Process stopped by user" in str(e):
            print("🛑 Stop flag detected. Exiting process_account.")
            raise
        else:
            print(f"⚠️ Process interrupted with error: {e}")
            raise


def pre_process_accounts_FWD(accounts_to_process, extension_page, site_page, profile_number):
    for account_index in accounts_to_process:
        print(f"\n🔄 Starting process for account [{account_index}]")
        if stop_event.is_set():
            print("🛑 Stop detected before processing account. Exiting loop.")
            break
        try:
            process_account_FWD(extension_page, site_page, account_index, profile_number)
            print(f"✅ Completed process for account {account_index}\n")
        except Exception as e:
            err_msg = str(e)
            if "Process stopped by user" in err_msg:
                print("🛑 Stop flag detected. Exiting.")
                break
            elif skip_event.is_set():
                print(f"⏭️ Account {account_index} skipped due to user command.")
                skip_event.clear()
                continue
            else:
                print(f"❌ Error processing account {account_index}: {e}")
                continue
        time.sleep(1)


def run6(profile_number: str, user_data_dir: str):
    user_input = input("\n📱 Enter account numbers = ")
    accounts_to_process = parse_account_input(user_input)
    print(f"💻 Accounts to process:{accounts_to_process}")
    print(f"🚀 Launching Edge with Profile {profile_number}\n")

    with sync_playwright() as p:
        context = p.chromium.launch_persistent_context(
            user_data_dir=user_data_dir,
            executable_path=browser_path,
            headless=False,
            args=[
                f"--disable-extensions-except={extension_path}",
                f"--load-extension={extension_path}",
            ]
        )

        Win_cl(context)

        extension_page = context.pages[0]
        safe_goto(extension_page, "chrome-extension://ejbalbakoplchlghecdalmeeeajnimhm/home.html")
        extension_page.set_viewport_size({"width": 1000, "height": 800})

        unlock_wallet(extension_page, password_extension)

        site_page = context.pages[1]
        site_page.set_viewport_size({"width": 1000, "height": 800})

        input_thread = threading.Thread(target=user_input_monitor, daemon=True)
        input_thread.start()

        try:
            pre_process_accounts_FWD(accounts_to_process, extension_page, site_page, profile_number)
        finally:
            print("👌 All done or stopped.")
            input(f"📌 Press Enter to close Profile {profile_number}")
            context.close()

##############################################################################################################################################

if __name__ == "__main__":
    choice = input("📂 Enter profile numbers = ").strip()
    if choice not in USER_PROFILES:
        print(f"❌ Invalid profile number '{choice}'. Please check USER_PROFILES.")
        exit()

    print("\nTASK with No."
          "\n 0 = Meta mesk"
          "\n 1 = Meta mesk Staking"
          "\n 2 = Meta mesk Farming"
          "\n 3 = Level 1 Data"
          "\n 4 = BNB Check"
          "\n 5 = PVC & BNB Data"
          "\n 6 = Farming Withdrawal")

    task_choice = input("\n⚙️ Which Task to run = ").strip()
    if task_choice == "0":
        task = run0
    elif task_choice == "1":
        task = run1
    elif task_choice == "2":
        task = run2
    elif task_choice == "3":
        task = run3
    elif task_choice == "4":
        task = run4
    elif task_choice == "5":
        task = run5
    elif task_choice == "6":
        task = run6

    else:
        print("❌ Invalid task choice. Please run again.")
        exit()
    thread = threading.Thread(target=task, args=(choice, USER_PROFILES[choice]))
    thread.start()
    thread.join()
    print("🆑 Browser closed.")